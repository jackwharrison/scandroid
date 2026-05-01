from flask import Flask, render_template, request, send_file, redirect, session, url_for, flash, jsonify, send_from_directory
import requests
from io import BytesIO
from config_loader import load_config, save_config
import json
import os
from config import ADMIN_USERNAME, ADMIN_PASSWORD, FSP_USERNAME, FSP_PASSWORD
from urllib.parse import quote
import subprocess
import zipfile
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A5, landscape
from reportlab.lib.units import cm
from reportlab.lib.utils import ImageReader
import qrcode
import csv
from flask_session import Session
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from openpyxl import load_workbook
from config_loader import load_display_config, save_display_config



# Register a Unicode-safe font
pdfmetrics.registerFont(TTFont("DejaVu", os.path.join("static", "fonts", "DejaVuSans.ttf")))


app = Flask(__name__)
app.secret_key = 'your_secret_key'
app.config["SESSION_TYPE"] = "filesystem"
app.config["SESSION_PERMANENT"] = False
Session(app)

@app.context_processor
def inject_national_society():
    config = load_config()
    return {"national_society": config.get("nationalSociety", "")}

def _make_qr_image(data, box_cm=3.0):
    """Return a Pillow image for the QR sized to box_cm × box_cm at 300dpi."""
    qr = qrcode.QRCode(
        version=None, error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=10, border=4
    )
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white").convert("RGB")

    # resize to cm at 300dpi (ReportLab draws images in points, we’ll scale when drawing)
    target_px = int((box_cm / 2.54) * 300)  # cm -> inches -> px
    img = img.resize((target_px, target_px))
    return img

def _draw_voucher(c, item, static_folder):
    """
    Draw one voucher on an A5 LANDSCAPE page matching the provided layout.
    Supports dynamic CSV fields.
    """
    width, height = landscape(A5)

    margin = 1.0 * cm
    inner_w = width - 2 * margin
    inner_h = height - 2 * margin

    # ---- BORDER -------------------------------------------------------------
    c.setLineWidth(0.8)
    c.rect(margin, margin, inner_w, inner_h)
    logo_height = 2.5 * cm       # height of left logo
    logo_top_y = height - margin - 1.0*cm   # move logos DOWN slightly (was -0.0)

    # LEFT LOGO (ns1)
    try:
        logo1 = ImageReader(os.path.join(static_folder, "ns1.png"))
        c.drawImage(
            logo1,
            margin,                          # <--- as far left as possible
            logo_top_y - logo_height,
            preserveAspectRatio=True,
            height=logo_height,
            mask='auto'
        )
        left_logo_bottom = logo_top_y - logo_height
        left_logo_top = logo_top_y
    except:
        left_logo_top = height - margin - 0.5*cm

    # RIGHT LOGO (ns2) – wide logo, scale by width
    try:
        logo2 = ImageReader(os.path.join(static_folder, "ns2.png"))
        max_width = 5.0 * cm

        img_w, img_h = logo2.getSize()
        scale = max_width / img_w
        scaled_height = img_h * scale

        c.drawImage(
            logo2,
            width - margin - max_width,
            logo_top_y - scaled_height,
            width=max_width,
            height=scaled_height,
            preserveAspectRatio=True,
            mask='auto'
        )
        right_logo_top = logo_top_y
    except:
        right_logo_top = height - margin - 0.5*cm

    # --- CENTER "Project" between logos (vertically aligned to logos) ---
    project_y = min(left_logo_top, right_logo_top) - 0.3*cm

    c.setFont("DejaVu", 16)
    c.drawCentredString(width/2, project_y, "Project")

    # underline "Project"
    text_width = c.stringWidth("Project", "DejaVu", 16)
    c.line(
        (width/2 - text_width/2),
        project_y - 0.08*cm,
        (width/2 + text_width/2),
        project_y - 0.08*cm
    )

    # --- MAIN TITLE just below "Project" ---
    title_y = project_y - 1.4*cm
    c.setFont("DejaVu", 22)
    c.drawCentredString(width/2, title_y, "CASH ON THE MOVE")

    # Subtitle lines
    c.setFont("DejaVu", 11)
    c.drawCentredString(width/2, title_y - 1.0*cm,
                        "Supporting people on the move especially those")
    c.drawCentredString(width/2, title_y - 1.6*cm,
                        "in vulnerable situations")
    # ---- QR CODE ------------------------------------------------------------
    qr_box_size = 5.0 * cm
    qr_x = margin + 0.6*cm
    qr_y = margin + 0.6*cm

    # Outer box
    c.setLineWidth(0.7)
    c.rect(qr_x - 0.1*cm, qr_y - 0.1*cm, qr_box_size + 0.2*cm, qr_box_size + 0.2*cm)

    # QR image
    refid = item.get("referenceid", "").strip()
    qr_img = _make_qr_image(refid, box_cm=3.0)
    c.drawInlineImage(qr_img, qr_x, qr_y, width=qr_box_size, height=qr_box_size)

    # *** Removed “Reference ID below QR” (as you requested) ***

    # ---- BENEFICIARY INFORMATION -------------------------------------------
    info_x = qr_x + qr_box_size + 2.0*cm
    info_y = qr_y + qr_box_size - 0.5*cm
    line_height = 0.75 * cm

    c.setFont("DejaVu", 12)

    # Formatting for dynamic keys
    def pretty_label(raw):
        raw = raw.replace("_", " ")
        raw = ''.join([' ' + ch if ch.isupper() else ch for ch in raw])
        raw = ' '.join(raw.split())
        return raw.title()

    # All dynamic fields except referenceId
    def is_reference_id(key: str) -> bool:
        k = key.strip().lower().replace("_", "").replace(" ", "")
        return ("ref" in k and "id" in k)
        

    # All dynamic fields except the reference ID (robust detection)
    fields_to_print = {
        k: v for k, v in item.items()
        if not is_reference_id(k) and v not in (None, "")
    }


    y = info_y
    for key, value in fields_to_print.items():
        label = pretty_label(key)
        c.setFont("DejaVu", 12)
        c.drawString(info_x, y, f"{label}: {value}")
        y -= line_height

    # ---- SMALL REFERENCE ID AT BOTTOM ------------------------------------
    refid = item.get("referenceid", "").strip()
    if refid:
        c.setFont("DejaVu", 8)
        c.drawString(
            margin + 0.1*cm,
            margin + 0.2*cm,
            f"Reference ID: {refid}"
        )


def generate_vouchers_pdf(rows, static_folder):
    """
    rows: list of dicts with keys: referenceId, name
    returns BytesIO of PDF
    """
    from io import BytesIO
    pdf_io = BytesIO()

    # Create a landscape A5 page
    c = canvas.Canvas(pdf_io, pagesize=landscape(A5))

    for r in rows:
        _draw_voucher(c, r, static_folder)
        c.showPage()

    c.save()
    pdf_io.seek(0)
    return pdf_io

translations = {
"en": {
    "title": "Beneficiary Information",
    "name": "Name",
    "dob": "Date of Birth",
    "photo": "Photo",
    "payment_approved": "Payment Approved",
    "payment_rejected": "Payment Rejected",
    "participant_withdraws": "Participant Withdraws",
    "language": "Language",
    "login": "Login for Red Cross Staff",
    "enter_password": "Enter Password",
    "submit": "Submit",
    "payment_status": "Payment Status",
    "confirm_person": "Confirm this is the correct person",
    "rejection_reason": "Reason for Rejection",
    "already_scanned": "This person has already been scanned and submitted to 121.",
    "success_message": "Payment successfully submitted.",
    "already_submitted_page": "This beneficiary's payment has already been submitted. If you need support, contact support@121.global",
    "config_title": "Configure Fields to Display",
    "field_key": "Kobo Photo Field",
    "label_en": "Label (EN)",
    "label_fr": "Label (FR)",
    "label_ar": "Label (AR)",
    "remove": "Remove",
    "add_field": "Add Field",
    "save": "Save",
    "saved_successfully": "Saved successfully",
    "failed_to_save": "Failed to save",
    "logout": "Logout",
    "config_system": "System Configuration",
    "config_display": "Configure Fields to Display",
    "fsp_login": "Login for FSP Admins",
    "program_select_title": "Select a Program",
    "back_to_programs": "Back to Programs",
    "program_select_subtitle": "Choose the program you want to work with",
    "fsp_sync_title": "Prepare to Scan",
    "sync_latest": "Sync Latest Records",
    "syncing": "Syncing...",
    "sync_error": "❌ Failed to sync. Please try again.",
    "sync_initial": "Click sync to see how many beneficiaries are ready for offline validation.",
    "sync_complete": "✅ {count} beneficiaries ready for offline validation.",
    "step1": "Step 1. Sync with 121",
    "step2": "Step 2. Save Latest Records",
    "step3": "Step 3. Scan QR Codes",
    "online": "Online",
    "offline": "Offline",
    "scan_title": "Scan QR",
    "back_to_dashboard": "Back",
    "scan_hint": "Point the camera at the QR code.",
    "start_camera": "Start camera",
    "waiting_to_start": "Waiting to start…",
    "requesting_camera": "Requesting camera… If prompted, tap Allow.",
    "camera_denied": "Camera permission denied or not available.",
    "scanning": "Scanning…",
    "starting_camera": "Starting camera…",
    "footer_dev": "Developed by 510 @ The Netherlands Red Cross",
    "footer_support": "If you need support contact jharrison@redcross.nl",
    "kobo_info": "Kobo Information",
    "kobo_token": "Kobo API Key",
    "asset_id": "Kobo Asset ID",
    "fsp_password": "Set Password for FSPs",
    "password": "Password",
    "encryption_settings": "Encryption Settings",
    "encryption_key": "Encryption Key",
    "encryption_warning": "Used to decrypt encrypted fields. If incorrect, offline validation will stop working.",
    "encryption_toggle_warning": "I understand changing this may break the system if incorrect.",
    "info_121": "121 Information",
    "url121": "121 URL",
    "username121": "121 Username",
    "program_id": "Select Program",
    "payment_id": "Payment ID",
    "column_to_match": "Field to Match for Payment (e.g., phoneNumber)",
    "column_to_match_info": "This field is selected in the Field Display Config page.",
    "use_for_matching": "Use for Matching Payments",
    "photo_config_title": "Photo Field Configuration",
    "enable_photo_field": "Enable photo field display",
    "home_question": "Who are you?",
    "home_admin": "Red Cross Staff",
    "home_fsp": "Financial Service Provider",
    "fsp_login": "Log in for Financial Service Provider",
    "step_4_generate": "📤 Step 4. Generate Payments to Send to 121",
    "payments_ready": "🔄 Payments ready to submit to 121:",
    "generate_csv": "Step 4. Generate Payments CSV",
    "download_csv": "⬇️ Download CSV",
    "step_5_send": "✅ Step 5. Send Payments to 121",
    "send_payments": "Step 5. Send payments to 121",
    "payment_submit_success": "✅ Payments submitted successfully!",
    "payment_submit_failed": "❌ Failed to submit",
    "voucher_generator": "Voucher Generator",
    "csv_hint": "Upload a CSV or Excel file with referenceId and any extra fields to print",
    "choose_csv": "Choose CSV or Excel file…",
    "upload_csv": "Upload CSV or Excel file",
    "download_vouchers": "Download vouchers (PDF)",
    "back": "Back to Dashboard",
    "choose_csv_alert": "Please choose a CSV first.",
    "upload_failed": "Upload failed",
    "voucher_ready_singular": "voucher ready to download",
    "voucher_ready_plural": "vouchers ready to download",
    "generate_vouchers": "Generate Vouchers",
    "invalid_qr_title": "Invalid QR Code",
    "invalid_qr_message": "This QR code cannot be used.",
    "checking_reason": "Checking reason…",
    "scan_next": "Scan next beneficiary",
    "go_home": "Go to homepage",
    "reason_used": "This QR code has already been used.",
    "reason_invalid": "The QR code is invalid or unrecognized.",
    "reason_no_record": "This QR code does not match any stored beneficiary.",
    "reason_database": "There was a problem reading offline data. Please try again.",
    "success_title": "Successfully submitted",
    "success_message": "You may now scan the next beneficiary.",
    "loading_counter": "Payments ready to push to 121: ...",
    "payments_ready": "Payments ready to push to 121:",
    "scan_next": "Scan next beneficiary",
    "go_home": "Finished scanning? Go to homepage",
    "payment_prep": "Prepare to Send Payments",
    "login_error": "Incorrect username or password.",
    "admin_dashboard_title": "Admin Dashboard",
    "admin_dashboard_subtitle": "Manage your Scandroid configuration and voucher tools.",
    "config_system_desc": "Edit API keys, endpoints, and system parameters.",
    "config_display_desc": "Configure which fields and details are shown to FSPs.",
    "generate_vouchers_desc": "Generate QR vouchers for printing and distribution.",
    "program_information": "Program Information",
    "credentials": "Credentials",
    "url_121": "121 URL",
    "username_121": "121 Username",
    "password_121": "121 Password",
    "field_name_121": "121 Field Name",
    "matching_field_title": "121 Field for Matching Payments",
    "matching_field_desc": "This field contains the unique ID from 121 used to match beneficiaries to payments.",
    "matching_field_placeholder": "e.g. phoneNumber",
    "wrong_credentials": "Incorrect username or password.",
    "api_unreachable": "Unable to reach the login server. Please try again.",
    "login_failed_generic": "Login failed. Please try again.",
    "go_back": "Go Back",
    "program_title": "Program Title",
    "connected_to": "Connected to form:",
    "form_owner": "Owner",
    "no_form_connected": "Unable to load form details",
    "kobo_connection": "Kobo Connection",
    "payment_amount": "Payment Amount",
}
,
"fr": {
    "title": "Informations sur le bénéficiaire",
    "name": "Nom",
    "dob": "Date de naissance",
    "photo": "Photo",
    "payment_approved": "Paiement approuvé",
    "payment_rejected": "Paiement refusé",
    "participant_withdraws": "Le participant se retire",
    "language": "Langue",
    "login": "Connexion pour le personnel de la Croix-Rouge",
    "enter_password": "Entrer le mot de passe",
    "submit": "Soumettre",
    "payment_status": "Statut du paiement",
    "confirm_person": "Confirmez que c'est la bonne personne",
    "rejection_reason": "Motif du refus",
    "already_scanned": "Cette personne a déjà été scannée et soumise à 121.",
    "success_message": "Paiement soumis avec succès.",
    "already_submitted_page": "Le paiement de ce bénéficiaire a déjà été soumis. Si vous avez besoin d'aide, contactez support@121.global",
    "config_title": "Configurer les champs à afficher",
    "field_key": "Clé de champ (depuis Kobo)",
    "label_en": "Libellé (EN)",
    "label_fr": "Libellé (FR)",
    "label_ar": "Libellé (AR)",
    "remove": "Supprimer",
    "add_field": "Ajouter un champ",
    "save": "Enregistrer",
    "saved_successfully": "Enregistré avec succès",
    "failed_to_save": "Échec de l'enregistrement",
    "logout": "Déconnexion",
    "config_system": "Configurer le système ",
    "config_display": "Congifuration des champs de vérification",
    "fsp_login": "Connexion pour les FSP",
    "program_select_title": "Sélectionner un programme",
    "back_to_programs": "Retour aux programmes",
    "program_select_subtitle": "Choisissez le programme avec lequel vous souhaitez travailler",
    "fsp_sync_title": "Synchroniser les enregistrements hors ligne",
    "sync_latest": "Synchroniser les derniers enregistrements",
    "syncing": "Synchronisation...",
    "sync_error": "❌ Échec de la synchronisation. Veuillez réessayer.",
    "sync_initial": "Cliquez sur synchroniser pour voir combien de bénéficiaires sont prêts pour la validation hors ligne.",
    "sync_complete": "✅ {count} bénéficiaires prêts pour la validation hors ligne.",
    "step1": "Étape 1. Synchroniser les derniers enregistrements",
    "step2": "Étape 2. Importer le cache hors ligne",
    "step3": "Étape 3. Scanner les codes QR",
    "online": "En ligne",
    "offline": "Hors ligne",
    "scan_title": "Scanner un QR",
    "back_to_dashboard": "Retour au tableau de bord",
    "scan_hint": "Pointez la caméra vers le code QR.",
    "start_camera": "Démarrer la caméra",
    "waiting_to_start": "En attente de démarrage…",
    "requesting_camera": "Demande d’accès à la caméra… Si demandé, touchez Autoriser.",
    "camera_denied": "Accès à la caméra refusé ou non disponible.",
    "scanning": "Analyse…",
    "starting_camera": "Démarrage de la caméra…",
    "footer_dev": "Développé par 510 @ La Croix-Rouge néerlandaise",
    "footer_support": "Pour toute assistance, contactez jharrison@redcross.nl",
    "kobo_info": "Informations Kobo",
    "kobo_token": "Clé API",
    "asset_id": "ID d'actif Kobo",
    "fsp_password": "Définir un mot de passe pour les FSP",
    "password": "Mot de passe",
    "encryption_settings": "Paramètres de chiffrement",
    "encryption_key": "Clé de chiffrement",
    "encryption_warning": "Utilisée pour déchiffrer les champs. Si elle est incorrecte, la validation hors ligne ne fonctionnera pas.",
    "encryption_toggle_warning": "Je comprends que changer cela pourrait casser le système si incorrect.",
    "info_121": "Informations 121",
    "url121": "URL 121",
    "username121": "Nom d'utilisateur 121",
    "program_id": "Sélectionner le programme",
    "payment_id": "ID de paiement",
    "column_to_match": "Champ à faire correspondre pour le paiement (ex. : phoneNumber)",
    "column_to_match_info": "Ce champ est sélectionné dans la page de configuration d'affichage.",
    "use_for_matching": "Utiliser pour le rapprochement des paiements",
    "photo_config_title": "Configuration du champ photo",
    "enable_photo_field": "Activer l'affichage du champ photo",
    "home_question": "Qui es-tu?",
    "home_admin": "Personnel de la Croix-Rouge",
    "home_fsp": "Prestataire de services financiers",
    "fsp_login": "Connexion pour le prestataire de services financiers",
    "step_4_generate": "📤 Étape 4. Générer les paiements à envoyer à 121",
    "payments_ready": "🔄 Paiements prêts à être soumis à 121 :",
    "generate_csv": "Étape 4. Générer un CSV",
    "download_csv": "⬇️ Télécharger le CSV",
    "step_5_send": "✅ Étape 5. Envoyer les paiements à 121",
    "send_payments": "Étape 5. Envoyer les paiements à 121",
    "payment_submit_success": "✅ Paiements envoyés avec succès !",
    "payment_submit_failed": "❌ Échec de l'envoi",
    "voucher_generator": "Générateur de bons",
    "csv_hint": "Téléchargez un fichier CSV ou Excel avec referenceId et d’autres champs à imprimer",
    "choose_csv": "Télécharger un fichier CSV ou Excel",
    "upload_csv": "Importer le fichier",
    "download_vouchers": "Télécharger les bons (PDF)",
    "back": "Retour au tableau de bord",
    "choose_csv_alert": "Veuillez d'abord choisir un fichier CSV.",
    "upload_failed": "Échec du téléversement",
    "voucher_ready_singular": "bon prêt à télécharger",
    "voucher_ready_plural": "bons prêts à télécharger",
    "generate_vouchers": "Générer les coupons",
    "invalid_qr_title": "Code QR invalide",
    "invalid_qr_message": "Ce code QR ne peut pas être utilisé.",
    "checking_reason": "Vérification de la raison…",
    "scan_next": "Scanner le bénéficiaire suivant",
    "go_home": "Aller à l'accueil",
    "reason_used": "Ce code QR a déjà été utilisé.",
    "reason_invalid": "Le code QR est invalide ou non reconnu.",
    "reason_no_record": "Aucun bénéficiaire correspondant n’a été trouvé.",
    "reason_database": "Problème de lecture des données hors ligne. Veuillez réessayer.",
    "success_title": "Soumis avec succès",
    "success_subtitle": "Vous pouvez maintenant scanner le bénéficiaire suivant.",
    "counter_loading": "Paiements prêts à envoyer à 121 : ...",
    "counter_label": "Paiements prêts à envoyer à 121 :",
    "scan_next": "Scanner le bénéficiaire suivant",
    "go_home": "Terminé le scan ? Aller à la page d’accueil",
    "payment_prep": "Préparer l'envoi des paiements",
    "login_error": "Nom d’utilisateur ou mot de passe incorrect.",
    "admin_dashboard_title": "Tableau de bord Admin",
    "admin_dashboard_subtitle": "Gérez la configuration de Scandroid et les outils de bons.",
    "config_system_desc": "Modifier les clés API, les points d'accès et les paramètres du système.",
    "config_display_desc": "Configurer les champs et informations affichés aux FSP.",
    "generate_vouchers_desc": "Générer des bons QR pour impression et distribution.",
    "program_information": "Informations du programme",
    "credentials": "Identifiants",
    "url_121": "URL 121",
    "username_121": "Nom d’utilisateur 121",
    "password_121": "Mot de passe 121",
    "field_name_121": "Nom du champ 121",
    "matching_field_title": "Champ 121 pour l’appariement des paiements",
    "matching_field_desc": "Ce champ contient l’identifiant unique de 121 utilisé pour faire correspondre les bénéficiaires aux paiements.",
    "matching_field_placeholder": "ex. phoneNumber",
    "wrong_credentials": "Nom d’utilisateur ou mot de passe incorrect.",
    "api_unreachable": "Impossible de joindre le serveur. Veuillez réessayer.",
    "login_failed_generic": "Échec de connexion. Veuillez réessayer.",
    "go_back": "Retour",
    "program_title": "Titre du programme",
    "connected_to": "Connecté au formulaire :",
    "form_owner": "Propriétaire",
    "no_form_connected": "Impossible de charger les détails du formulaire",
    "kobo_connection": "Connexion Kobo",
    "payment_amount": "Montant du paiement",
}
,
"ar": {
    "title": "معلومات المستفيد",
    "name": "الاسم",
    "dob": "تاريخ الميلاد",
    "photo": "صورة",
    "payment_approved": "تمت الموافقة على الدفع",
    "payment_rejected": "تم رفض الدفع",
    "participant_withdraws": "انسحب المستفيد",
    "language": "اللغة",
    "login": "تسجيل دخول لموظفي الهلال الأحمر",
    "enter_password": "أدخل كلمة المرور",
    "submit": "إرسال",
    "payment_status": "حالة الدفع",
    "confirm_person": "تأكيد أن هذه هي الشخص الصحيح",
    "rejection_reason": "سبب الرفض",
    "already_scanned": "تم بالفعل مسح هذا الشخص وإرساله إلى 121.",
    "success_message": "تم إرسال الدفع بنجاح.",
    "already_submitted_page": "تم بالفعل إرسال دفعة هذا المستفيد. إذا كنت بحاجة إلى الدعم، فاتصل بـ support@121.global",
    "config_title": "تكوين الحقول المعروضة",
    "field_key": "مفتاح الحقل (من كوبا)",
    "label_en": "التسمية (EN)",
    "label_fr": "التسمية (FR)",
    "label_ar": "التسمية (AR)",
    "remove": "إزالة",
    "add_field": "إضافة حقل",
    "save": "حفظ",
    "saved_successfully": "تم الحفظ بنجاح",
    "failed_to_save": "فشل الحفظ",
    "logout": "تسجيل الخروج",
    "config_system": "إعدادات النظام",
    "config_display": "تكوين الحقول المعروضة",
    "fsp_login": "تسجيل الدخول لمزودي الخدمات المالية",
    "program_select_title": "اختر برنامجاً",
    "back_to_programs": "العودة إلى البرامج",
    "program_select_subtitle": "اختر البرنامج الذي تريد العمل معه",
    "fsp_sync_title": "📥 مزوّد الخدمة: مزامنة السجلات غير المتصلة",
    "sync_latest": "مزامنة أحدث السجلات",
    "syncing": "جارٍ المزامنة...",
    "sync_error": "❌ فشلت عملية المزامنة. يرجى المحاولة مرة أخرى.",
    "sync_initial": "انقر على المزامنة لمعرفة عدد المستفيدين الجاهزين للتحقق دون اتصال.",
    "sync_complete": "✅ {count} مستفيدين جاهزين للتحقق دون اتصال.",
    "step1": "الخطوة 1. مزامنة أحدث السجلات",
    "step2": "الخطوة 2. استيراد ذاكرة التخزين المؤقت دون اتصال",
    "step3": "الخطوة 3. مسح رموز QR",
    "online": "متصل",
    "offline": "غير متصل",
    "scan_title": "مسح رمز QR",
    "back_to_dashboard": "العودة إلى لوحة التحكم",
    "scan_hint": "وجّه الكاميرا نحو رمز QR.",
    "start_camera": "بدء تشغيل الكاميرا",
    "waiting_to_start": "بانتظار البدء…",
    "requesting_camera": "جارٍ طلب تشغيل الكاميرا… إذا طُلِب منك ذلك، اضغط سماح.",
    "camera_denied": "تم رفض إذن الكاميرا أو أنها غير متاحة.",
    "scanning": "جارٍ المسح…",
    "starting_camera": "جارٍ بدء تشغيل الكاميرا…",
    "footer_dev": "تم التطوير بواسطة 510 @ الصليب الأحمر الهولندي",
    "footer_support": "إذا كنت بحاجة إلى الدعم، تواصل مع jharrison@redcross.nl",
    "kobo_info": "معلومات كوبا",
    "kobo_token": "رمز كوبا",
    "asset_id": "معرف الأصول في كوبا",
    "fsp_password": "تعيين كلمة مرور لـ FSP",
    "password": "كلمة المرور",
    "encryption_settings": "إعدادات التشفير",
    "encryption_key": "مفتاح التشفير",
    "encryption_warning": "يُستخدم لفك تشفير الحقول المشفرة. إذا كان غير صحيح، فلن تعمل التحقق بدون اتصال.",
    "encryption_toggle_warning": "أفهم أن التغيير هنا قد يؤدي إلى تعطل النظام إذا كان غير صحيح.",
    "info_121": "معلومات 121",
    "url121": "رابط 121",
    "username121": "اسم مستخدم 121",
    "payment_id": "معرف الدفع",
    "column_to_match": "الحقل المطابق للدفع (مثل رقم الهاتف)",
    "column_to_match_info": "يتم تحديد هذا الحقل في صفحة إعداد عرض الحقول.",
    "use_for_matching": "استخدم لمطابقة المدفوعات",
    "photo_config_title": "إعدادات حقل الصورة",
    "enable_photo_field": "تفعيل عرض حقل الصورة",
    "home_question": "من أنت؟",
    "home_admin": "موظفو الصليب الأحمر",
    "home_fsp": "مزود خدمات مالية",
    "fsp_login": "تسجيل الدخول لمزود الخدمة المالية",
    "step_4_generate": "📤 الخطوة 4. إنشاء الدفعات لإرسالها إلى 121",
    "payments_ready": "🔄 الدفعات الجاهزة للإرسال إلى 121:",
    "generate_csv": "إنشاء ملف CSV",
    "download_csv": "⬇️ تحميل ملف CSV",
    "step_5_send": "✅ الخطوة 5. إرسال الدفعات إلى 121",
    "send_payments": "إرسال الدفعات",
    "payment_submit_success": "✅ تم إرسال المدفوعات بنجاح!",
    "payment_submit_failed": "❌ فشل الإرسال",
    "voucher_generator": "مولّد القسائم",
    "csv_hint": "قم بتحميل ملف CSV يحتوي على referenceId وأي حقول إضافية للطباعة",
    "choose_csv": "اختر ملف CSV/Excel…",
    "upload_csv": "تحميل ملف Excel/CSV",
    "download_vouchers": "تنزيل القسائم (PDF)",
    "back": "العودة إلى لوحة التحكم",
    "choose_csv_alert": "يرجى اختيار ملف CSV أولاً.",
    "upload_failed": "فشل الرفع",
    "voucher_ready_singular": "قسيمة جاهزة للتنزيل",
    "voucher_ready_plural": "قسائم جاهزة للتنزيل",
    "generate_vouchers": "إصدار القسائم",
    "invalid_qr_title": "رمز QR غير صالح",
    "invalid_qr_message": "لا يمكن استخدام رمز QR هذا.",
    "checking_reason": "جارٍ التحقق من السبب…",
    "scan_next": "مسح المستفيد التالي",
    "go_home": "العودة إلى الصفحة الرئيسية",
    "reason_used": "تم استخدام رمز QR هذا سابقًا.",
    "reason_invalid": "رمز QR غير صالح أو غير معروف.",
    "reason_no_record": "لا يوجد أي مستفيد مطابق لهذا الرمز.",
    "reason_database": "حدثت مشكلة في قراءة البيانات دون اتصال. حاول مرة أخرى.",
    "success_title": "تم الإرسال بنجاح",
    "success_subtitle": "يمكنك الآن مسح المستفيد التالي.",
    "counter_loading": "المدفوعات الجاهزة للإرسال إلى 121: ...",
    "counter_label": "المدفوعات الجاهزة للإرسال إلى 121:",
    "scan_next": "مسح المستفيد التالي",
    "go_home": "هل انتهيت من المسح؟ اذهب إلى الصفحة الرئيسية",
    "payment_prep": "Prepare to send payments",
    "login_error": "اسم المستخدم أو كلمة المرور غير صحيحة.",
    "admin_dashboard_title": "لوحة تحكم المسؤول",
    "admin_dashboard_subtitle": "إدارة إعدادات Scandroid وأدوات القسائم",
    "config_system_desc": "تعديل مفاتيح API ونقاط النهاية ومعلمات النظام.",
    "config_display_desc":"تكوين الحقول والمعلومات التي يتم عرضها لمقدمي الخدمات المالية.",
    "generate_vouchers_desc": "إنشاء قسائم QR للطباعة والتوزيع.",
    "program_information": "معلومات البرنامج",
    "credentials": "بيانات تسجيل الدخول",
    "url_121": "رابط 121",
    "program_id": "اختر البرنامج",
    "username_121": "اسم المستخدم 121",
    "password_121": "كلمة مرور 121",
    "field_name_121": "اسم الحقل في 121",
    "matching_field_title": "حقل 121 لمطابقة الدفعات",
    "matching_field_desc": "يحتوي هذا الحقل على المعرّف الفريد من نظام 121 المستخدم لمطابقة المستفيدين مع الدفعات.",
    "matching_field_placeholder": "مثال: phoneNumber",
    "wrong_credentials": "اسم المستخدم أو كلمة المرور غير صحيحة.",
    "api_unreachable": "تعذّر الاتصال بالخادم. يرجى المحاولة مرة أخرى.",
    "login_failed_generic": "فشل تسجيل الدخول. يرجى المحاولة مرة أخرى.",
    "go_back": "عودة",
    "program_title": "عنوان البرنامج",
    "connected_to": "متصل بالنموذج:",
    "form_owner": "المالك",
    "no_form_connected": "تعذّر تحميل تفاصيل النموذج",
    "kobo_connection": "اتصال كوبا",
    "payment_amount": "مبلغ الدفع",
    }
}

@app.route("/instance-static/<filename>")
def instance_static(filename):
    context = os.getenv("SCANDROID_CONTEXT", "local")
    azure_path = f"/home/site/configs/{context}/static"

    # Azure instance-specific logos
    instance_file = os.path.join(azure_path, filename)
    if os.path.exists(instance_file):
        return send_from_directory(azure_path, filename)

    # Local fallback for development
    return send_from_directory("static", filename)

@app.route("/")
def landing_page():
    lang = request.args.get("lang", "en")
    return render_template("home.html", lang=lang, t=translations.get(lang, translations["en"]))

@app.route("/admin-login", methods=["GET", "POST"])
def admin_login():
    # language handling
    lang = request.args.get("lang") or request.form.get("lang") or session.get("lang", "en")
    session["lang"] = lang
    t = translations.get(lang, translations["en"])

    # GET → show login page
    if request.method == "GET":
        return render_template("admin_login.html", lang=lang, t=t, error=None)

    # POST → authenticate against 121 API using system_config.json
    username = request.form.get("username")
    password = request.form.get("password")

    config = load_config()
    base_url = config.get("url121")
    if not base_url:
        return render_template(
            "admin_login.html",
            lang=lang, t=t,
            error="❌ Missing url121 in system configuration."
        )

    login_url = f"{base_url}/api/users/login"
    login_payload = {"username": username, "password": password}

    try:
        res = requests.post(login_url, json=login_payload, timeout=10)

    except Exception:
        # API unreachable
        return render_template(
            "admin_login.html",
            lang=lang, t=t,
            error=t.get("api_unreachable", "Unable to reach login server.")
        )

    # ✔ Success
    if res.status_code == 201:
        session["admin_logged_in"] = True
        session["admin_username"] = username
        return redirect(url_for("admin_dashboard", lang=lang))

    # ❌ Wrong username/password
    if res.status_code in (400, 401):
        return render_template(
            "admin_login.html",
            lang=lang,
            t=t,
            error=t.get("wrong_credentials", "Incorrect username or password.")
        )

    # ❌ Any other response
    return render_template(
        "admin_login.html",
        lang=lang,
        t=t,
        error=f"Login failed ({res.status_code})."
    )



@app.route("/admin-dashboard")
def admin_dashboard():
    if not session.get("admin_logged_in"):
        return redirect(url_for("admin_login", lang=request.args.get("lang", "en")))

    lang = request.args.get("lang", "en")
    t = translations.get(lang, translations["en"])

    # Load title from system_config.json
    config = load_config()
    program_title = config.get("programTitle", "")

    return render_template(
        "admin_dashboard.html",
        program_title=program_title,
        lang=lang,
        t=t,
        username=session.get("admin_username")
    )

@app.route("/admin-logout")
def admin_logout():
    lang = request.args.get("lang", "en")
    session.pop("admin_logged_in", None)
    return redirect(url_for("admin_login", lang=lang))


from flask import request, session, redirect, url_for, flash, render_template, jsonify
import requests
from requests.auth import HTTPBasicAuth
import json

@app.route("/system-config", methods=["GET", "POST"])
def system_config():
    if not session.get("admin_logged_in"):
        return redirect(url_for("admin_login", lang=request.args.get("lang", "en")))

    lang = request.args.get("lang", "en")
    t = translations.get(lang, translations["en"])
    config = load_config()

    # ============================================================
    # POST: SAVE CONFIG
    # ============================================================
    if request.method == "POST":
        updated = config.copy()

        # ---- Global fields ----
        # Always save these from the form
        for key in ["KOBO_SERVER", "url121"]:
            updated[key] = request.form.get(key, "").strip()

        # Only overwrite if a new value was submitted
        for key in ["KOBO_TOKEN", "username121", "password121", "ENCRYPTION_KEY"]:
            submitted = request.form.get(key, "").strip()
            if submitted:
                updated[key] = submitted
            # else: keep existing value from config.copy()

        # ---- Program mappings ----
        programs = []
        program_ids = request.form.getlist("PROGRAMS[][programId]")
        asset_ids = request.form.getlist("PROGRAMS[][koboAssetId]")

        kobo_server = updated.get("KOBO_SERVER")
        kobo_token = updated.get("KOBO_TOKEN")

        for pid, asset_id in zip(program_ids, asset_ids):
            if not pid or not asset_id:
                continue

            entry = {
                "programId": int(pid),
                "koboAssetId": asset_id.strip()
            }

            # ---- Validate Kobo asset ----
            try:
                r = requests.get(
                    f"{kobo_server}/api/v2/assets/{asset_id}/?format=json",
                    headers={"Authorization": f"Token {kobo_token}"},
                    timeout=10
                )
                if r.status_code == 200:
                    j = r.json()
                    entry["koboFormName"] = j.get("name")
                    entry["koboFormOwner"] = j.get("owner__username")
            except Exception as e:
                print("Kobo validation error:", e)

            programs.append(entry)

        updated["PROGRAMS"] = programs

        save_config(updated)
        flash(t["saved_successfully"])
        return redirect(url_for("system_config", lang=lang))

    # ============================================================
    # GET: LOAD DATA FOR UI
    # ============================================================
    url121 = config.get("url121")
    username121 = config.get("username121")
    password121 = config.get("password121")

    token = None
    program_ids = []
    program_options = []

    # ---- Login to 121 ----
    if url121 and username121 and password121:
        try:
            r = requests.post(
                f"{url121}/api/users/login",
                json={"username": username121, "password": password121}
            )
            if r.status_code == 201:
                j = r.json()
                token = j.get("access_token_general")
                program_ids = [int(pid) for pid in j.get("permissions", {}).keys()]
        except Exception as e:
            print("121 login failed:", e)

    # ---- Load program titles ----
    if token:
        for pid in program_ids:
            try:
                r = requests.get(
                    f"{url121}/api/programs/{pid}",
                    cookies={"access_token_general": token}
                )
                if r.status_code == 200:
                    pdata = r.json()
                    titles = pdata.get("titlePortal", {})
                    title = titles.get(lang) or next(iter(titles.values()), f"Program {pid}")
                    program_options.append({"id": pid, "title": title})
            except Exception as e:
                print(f"Program load failed ({pid}):", e)

    # Existing mappings for UI
    program_mappings = config.get("PROGRAMS", [])

    return render_template(
        "system_config.html",
        config=config,
        program_options=program_options,
        program_mappings=program_mappings,
        username=session.get("admin_username"),
        lang=lang,
        t=t
    )



@app.route("/api/program-attributes/<int:program_id>")
def api_program_attributes(program_id):
    if not session.get("admin_logged_in"):
        return jsonify({"error": "Unauthorized"}), 401

    system_config = load_config()
    url121 = system_config.get("url121")

    if not url121:
        return jsonify({"attributes": [], "kobo_image_fields": []})

    attributes = []
    kobo_image_fields = []

    try:
        login_resp = requests.post(
            f"{url121}/api/users/login",
            json={
                "username": system_config.get("username121", ""),
                "password": system_config.get("password121", "")
            },
            timeout=10
        )

        if login_resp.status_code == 201:
            token = login_resp.json().get("access_token_general")
            cookies = {"access_token_general": token}

            # Registration attributes
            r = requests.get(
                f"{url121}/api/programs/{program_id}",
                cookies=cookies,
                timeout=10
            )
            if r.status_code == 200:
                for attr in r.json().get("programRegistrationAttributes", []):
                    name = attr.get("name")
                    if not name:
                        continue
                    labels = attr.get("label") or {}
                    label = labels.get("en") or next(iter(labels.values()), name)
                    attributes.append({"name": name, "label": label})

    except Exception as e:
        print(f"[api_program_attributes] Error: {e}")

    # Kobo image fields for this program
    try:
        programs = system_config.get("PROGRAMS", [])
        program = next((p for p in programs if str(p.get("programId")) == str(program_id)), None)
        if program:
            asset_id = program.get("koboAssetId")
            kobo_token = system_config.get("KOBO_TOKEN")
            kobo_server = system_config.get("KOBO_SERVER", "https://kobo.ifrc.org")

            if asset_id and kobo_token:
                r = requests.get(
                    f"{kobo_server}/api/v2/assets/{asset_id}/?format=json",
                    headers={"Authorization": f"Token {kobo_token}"},
                    timeout=10
                )
                if r.status_code == 200:
                    survey = r.json().get("content", {}).get("survey", [])
                    for item in survey:
                        if not isinstance(item, dict) or item.get("type") != "image":
                            continue
                        raw_label = item.get("label")
                        if isinstance(raw_label, list) and raw_label:
                            label = raw_label[0] if isinstance(raw_label[0], str) else next(iter(raw_label[0].values()), "")
                        elif isinstance(raw_label, dict):
                            label = raw_label.get("en") or next(iter(raw_label.values()), "")
                        else:
                            label = str(raw_label or "")
                        xpath = item.get("$xpath", "").replace("/data/", "").replace("data/", "").strip("/")
                        name = xpath or item.get("name", "")
                        if name:
                            kobo_image_fields.append({"name": name, "label": label})
    except Exception as e:
        print(f"[api_program_attributes] Kobo error: {e}")

    return jsonify({"attributes": attributes, "kobo_image_fields": kobo_image_fields})

@app.route("/config", methods=["GET", "POST"])
def config_page():
    if not session.get("admin_logged_in"):
        return redirect(url_for("admin_login", lang=request.args.get("lang", "en")))

    lang = request.args.get("lang", "en")
    username = session.get("admin_username", "Admin")

    # ------------------------------------------------------
    # POST (SAVE DISPLAY CONFIG – PER PROGRAM)
    # ------------------------------------------------------
    if request.method == "POST":
        config_data = request.get_json()

        program_id = str(config_data.pop("programId"))
        config_data.pop("COLUMN_TO_MATCH", None)

        try:
            full_config = load_display_config()
            if "programs" not in full_config:
                full_config["programs"] = {}

            full_config["programs"][program_id] = config_data
            save_display_config(full_config)

            return jsonify({"success": True})
        except Exception as e:
            return jsonify({"success": False, "error": str(e)}), 500

    # ------------------------------------------------------
    # GET (LOAD PAGE)
    # ------------------------------------------------------

    # Load display config
    try:
        config_data = load_display_config()
    except Exception:
        config_data = {}

    # Load system config
    try:
        system_config = load_config()
    except Exception:
        system_config = {}

    programs_raw = system_config.get("PROGRAMS", [])
    programs = []          # UI list (id + title)
    program_lookup = {}    # Logic lookup (full object)

    url121 = system_config.get("url121")

    # ------------------------------------------------------
    # Resolve program titles from 121
    # ------------------------------------------------------
    if url121 and programs_raw:
        try:
            login_resp = requests.post(
                f"{url121}/api/users/login",
                json={
                    "username": system_config.get("username121", ""),
                    "password": system_config.get("password121", "")
                }
            )

            if login_resp.status_code == 201:
                token = login_resp.json().get("access_token_general")
                cookies = {"access_token_general": token}

                for p in programs_raw:
                    pid = p.get("programId")
                    title = str(pid)

                    try:
                        r = requests.get(
                            f"{url121}/api/programs/{pid}",
                            cookies=cookies,
                            timeout=10
                        )
                        if r.status_code == 200:
                            titles = r.json().get("titlePortal", {})
                            title = titles.get("en") or next(iter(titles.values()), title)
                    except Exception:
                        pass

                    programs.append({
                        "id": str(pid),
                        "title": title
                    })
                    program_lookup[str(pid)] = p

        except Exception as e:
            print("Program title lookup failed:", e)

    # Safety fallback
    if not programs:
        for p in programs_raw:
            pid = str(p.get("programId"))
            programs.append({"id": pid, "title": pid})
            program_lookup[pid] = p

    # ------------------------------------------------------
    # Active program selection
    # ------------------------------------------------------
    active_program_id = request.args.get("programId")
    if not active_program_id and programs:
        active_program_id = programs[0]["id"]
    active_program_id = str(active_program_id) if active_program_id else None

    # ------------------------------------------------------
    # Backward compatibility (single → multi)
    # ------------------------------------------------------
    if "programs" not in config_data:
        default_program_id = active_program_id or "default"
        config_data = {
            "programs": {
                default_program_id: {
                    "fields": config_data.get("fields", []),
                    "photo": config_data.get(
                        "photo",
                        {
                            "enabled": True,
                            "labels": {"en": "Photo", "fr": "Photo", "ar": "صورة"}
                        }
                    )
                }
            }
        }

    # ------------------------------------------------------
    # Resolve active program Kobo + 121 config
    # ------------------------------------------------------
    program_id = None
    asset_id = None

    for prog in system_config.get("PROGRAMS", []):
        if str(prog.get("programId")) == str(active_program_id):
            program_id = prog.get("programId")
            asset_id = prog.get("koboAssetId")
            break

    program_title = system_config.get("programTitle", "")
    column_to_match_121 = None
    allowed_attributes = []
    kobo_image_fields = []

    # ------------------------------------------------------
    # 121 PROGRAM ATTRIBUTES
    # ------------------------------------------------------
    if url121 and program_id:
        try:
            login_resp = requests.post(
                f"{url121}/api/users/login",
                json={
                    "username": system_config.get("username121", ""),
                    "password": system_config.get("password121", "")
                }
            )

            if login_resp.status_code == 201:
                token = login_resp.json().get("access_token_general")
                cookies = {"access_token_general": token}

                # Column to match
                try:
                    r = requests.get(
                        f"{url121}/api/programs/{program_id}/fsp-configurations",
                        cookies=cookies,
                        timeout=10
                    )
                    if r.status_code == 200:
                        for fsp in r.json():
                            for prop in fsp.get("properties", []):
                                if prop.get("name") == "columnToMatch":
                                    column_to_match_121 = prop.get("value")
                                    break
                except Exception:
                    pass

                # Registration attributes
                try:
                    r = requests.get(
                        f"{url121}/api/programs/{program_id}",
                        cookies=cookies,
                        timeout=10
                    )
                    if r.status_code == 200:
                        for attr in r.json().get("programRegistrationAttributes", []):
                            name = attr.get("name")
                            if not name:
                                continue

                            labels = attr.get("label") or {}
                            label = labels.get("en") or next(iter(labels.values()), name)

                            allowed_attributes.append({
                                "name": name,
                                "label": label
                            })
                except Exception:
                    pass

        except Exception as e:
            print("121 lookup failed:", e)

    # ------------------------------------------------------
    # Strict mode – drop invalid fields per program
    # ------------------------------------------------------
    if allowed_attributes:
        allowed_names = {a["name"] for a in allowed_attributes}
        for pdata in config_data.get("programs", {}).values():
            pdata["fields"] = [
                f for f in pdata.get("fields", [])
                if f.get("key") in allowed_names
            ]

    # ------------------------------------------------------
    # Kobo image fields
    # ------------------------------------------------------
    def get_full_kobo_path(item):
        if not isinstance(item, dict):
            return None

        if "$xpath" in item:
            return item["$xpath"].replace("/data/", "").replace("data/", "").strip("/")

        parts = []
        current = item
        while isinstance(current, dict):
            if current.get("name"):
                parts.append(current["name"])
            current = current.get("parent")

        return "/".join(reversed(parts)) if parts else None

    try:
        token = system_config.get("KOBO_TOKEN")
        kobo_server = system_config.get("KOBO_SERVER", "https://kobo.ifrc.org")

        if token and asset_id:
            r = requests.get(
                f"{kobo_server}/api/v2/assets/{asset_id}/?format=json",
                headers={"Authorization": f"Token {token}"},
                timeout=10
            )

            if r.status_code == 200:
                survey = r.json().get("content", {}).get("survey", [])
                for item in survey:
                    if not isinstance(item, dict):
                        continue

                    if item.get("type") == "image":
                        raw_label = item.get("label")

                        if isinstance(raw_label, dict):
                            label = (
                                raw_label.get("English")
                                or raw_label.get("en")
                                or next(iter(raw_label.values()), get_full_kobo_path(item))
                            )

                        elif isinstance(raw_label, list) and raw_label:
                            first = raw_label[0]
                            if isinstance(first, dict):
                                label = (
                                    first.get("English")
                                    or first.get("en")
                                    or next(iter(first.values()), get_full_kobo_path(item))
                                )
                            else:
                                label = str(first)

                        else:
                            label = get_full_kobo_path(item)

                        path = get_full_kobo_path(item)
                        if path:
                            kobo_image_fields.append({
                                "name": path,
                                "label": label
                            })
    except Exception as e:
        print("Kobo lookup failed:", e)

    # ------------------------------------------------------
    # Strict mode – image field per program
    # ------------------------------------------------------
    if kobo_image_fields:
        valid = {i["name"] for i in kobo_image_fields}
        for pdata in config_data.get("programs", {}).values():
            if pdata.get("photo", {}).get("field_name") not in valid:
                pdata.setdefault("photo", {})["field_name"] = ""

    # ------------------------------------------------------
    # Persist COLUMN_TO_MATCH
    # ------------------------------------------------------
    if column_to_match_121 and active_program_id:
        per_program = system_config.setdefault("COLUMN_TO_MATCH_PER_PROGRAM", {})
        per_program[str(active_program_id)] = column_to_match_121
        try:
            save_config(system_config)
        except Exception:
            pass

    # ------------------------------------------------------
    # RENDER
    # ------------------------------------------------------
    return render_template(
        "config.html",
        full_config=config_data,
        programs=programs,
        active_program_id=active_program_id,
        system_config=system_config,
        allowed_attributes=allowed_attributes,
        kobo_image_fields=kobo_image_fields,
        column_to_match_121=column_to_match_121 or system_config.get("COLUMN_TO_MATCH"),
        lang=lang,
        username=username,
        program_title=program_title,
        t=translations.get(lang, translations["en"]),
    )


@app.route("/logout")
def logout():
    lang = request.args.get("lang", "en")
    session.clear()
    return redirect(url_for("login", lang=lang))


@app.route("/fsp-login", methods=["GET", "POST"])
def fsp_login():
    lang = request.args.get("lang", "en")
    t = translations.get(lang, translations["en"])
    error = None

    config = load_config()
    base_url = config.get("url121")
    if not base_url:
        return render_template("fsp_login.html", lang=lang, t=t, error="❌ Missing url121")

    login_url = f"{base_url}/api/users/login"

    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")

        try:
            res = requests.post(
                login_url,
                json={"username": username, "password": password},
                timeout=8
            )

            if res.status_code == 201:
                session["fsp_logged_in"] = True
                session["fsp_username"] = username
                return redirect(url_for("fsp_program_selector"))

            elif res.status_code in (400, 401):
                error = t["login_error"]

            else:
                error = f"Login failed ({res.status_code})."

        except Exception:
            error = t["login_error"]

    return render_template("fsp_login.html", lang=lang, t=t, error=error)

@app.route("/fsp-programs")
def fsp_program_selector():
    # ------------------------------------------------------
    # Auth guard
    # ------------------------------------------------------
    if not session.get("fsp_logged_in"):
        return redirect(url_for("fsp_login"))

    # ------------------------------------------------------
    # Language + translations (MISSING BEFORE)
    # ------------------------------------------------------
    lang = request.args.get("lang", "en")
    t = translations.get(lang, translations["en"])
    username = session.get("fsp_username")

    # ------------------------------------------------------
    # Always initialise programs (CRITICAL FIX)
    # ------------------------------------------------------
    programs = []

    # ------------------------------------------------------
    # Load system config
    # ------------------------------------------------------
    try:
        system_config = load_config()
    except Exception:
        system_config = {}

    programs_raw = system_config.get("PROGRAMS", [])
    url121 = system_config.get("url121")

    # ------------------------------------------------------
    # Resolve program titles from 121
    # ------------------------------------------------------
    if url121 and programs_raw:
        try:
            login_resp = requests.post(
                f"{url121}/api/users/login",
                json={
                    "username": system_config.get("username121", ""),
                    "password": system_config.get("password121", "")
                },
                timeout=10
            )

            if login_resp.status_code == 201:
                token = login_resp.json().get("access_token_general")
                cookies = {"access_token_general": token}

                for p in programs_raw:
                    pid = p.get("programId")
                    title = str(pid)

                    try:
                        r = requests.get(
                            f"{url121}/api/programs/{pid}",
                            cookies=cookies,
                            timeout=10
                        )
                        if r.status_code == 200:
                            titles = r.json().get("titlePortal", {})
                            title = titles.get("en") or next(iter(titles.values()), title)
                    except Exception:
                        pass

                    programs.append({
                        "id": str(pid),
                        "title": title
                    })

        except Exception as e:
            print("Program lookup failed:", e)

    # ------------------------------------------------------
    # Fallback: titles = program IDs
    # ------------------------------------------------------
    if not programs:
        programs = [
            {
                "id": str(p.get("programId")),
                "title": str(p.get("programId"))
            }
            for p in programs_raw
        ]

    # ------------------------------------------------------
    # Render selector page
    # ------------------------------------------------------
    return render_template(
        "fsp_programs.html",
        programs=programs,
        lang=lang,
        t=t,
        username=username,
        program_title="Program Selector"
    )
@app.route("/select-program/<program_id>")
def select_program(program_id):
    if not session.get("fsp_logged_in"):
        return redirect(url_for("fsp_login"))

    # keep language if present
    lang = request.args.get("lang", "en")

    # store selection in session
    session["fsp_program_id"] = str(program_id)

    # redirect into the admin page WITH query param
    return redirect(url_for("fsp_admin", program_id=str(program_id), lang=lang))

@app.route("/fsp-admin")
def fsp_admin():
    # --- auth ---
    if not session.get("fsp_logged_in"):
        return redirect(url_for("fsp_login"))

    # --- language ---
    lang = request.args.get("lang", "en")
    t = translations.get(lang, translations["en"])

    # --- program resolution (URL first, then session) ---
    program_id = request.args.get("program_id") or session.get("fsp_program_id")
    if not program_id:
        return redirect(url_for("fsp_program_selector", lang=lang))

    program_id = str(program_id)

    # --- load configs ---
    system_config = load_config()
    _full_display = load_display_config()
    # Pass full display config to template so it can be stored in IndexedDB
    display_config = _full_display

    column_to_match = get_column_to_match(program_id) or system_config.get("COLUMN_TO_MATCH")
    programs = system_config.get("PROGRAMS", [])

    # --- resolve program ---
    program = next(
        (p for p in programs if str(p.get("programId")) == program_id),
        None
    )

    if not program:
        return redirect(url_for("fsp_program_selector", lang=lang))

    # --- fetch 121 program title ---
    program_title = f"Program {program_id}"
    url121 = system_config.get("url121")
    if url121:
        try:
            login_resp = requests.post(
                f"{url121}/api/users/login",
                json={
                    "username": system_config.get("username121", ""),
                    "password": system_config.get("password121", "")
                },
                timeout=8
            )
            if login_resp.status_code == 201:
                token = login_resp.json().get("access_token_general")
                r = requests.get(
                    f"{url121}/api/programs/{program_id}",
                    cookies={"access_token_general": token},
                    timeout=8
                )
                if r.status_code == 200:
                    titles = r.json().get("titlePortal", {})
                    program_title = titles.get(lang) or next(iter(titles.values()), program_title)
        except Exception as e:
            print(f"[fsp_admin] Failed to fetch 121 program title: {e}")

    username = session.get("fsp_username")

    # --- IMPORTANT: persist program for later routes ---
    session["fsp_program_id"] = program_id

    return render_template(
        "fsp_admin.html",
        COLUMN_TO_MATCH=column_to_match,
        display_config=display_config,
        lang=lang,
        t=t,
        config=system_config,
        program_title=program_title,
        program_id=program_id,   # ✅ this feeds ACTIVE_PROGRAM_ID in JS
        username=username
    )


@app.route("/sync-fsp")
def sync_fsp():
    import subprocess
    import os

    # 🔴 get selected program from session
    program_id = session.get("fsp_program_id")
    if not program_id:
        return jsonify({
            "success": False,
            "message": "❌ No program selected"
        })

    env = os.environ.copy()
    env["PROGRAM_ID"] = str(program_id)   # 🔑 THIS IS THE FIX

    try:
        result = subprocess.run(
            ["python", "offline_sync.py"],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            env=env                      # 🔑 PASS ENV
        )

        print("\n[DEBUG] STDOUT:\n", result.stdout)
        print("\n[DEBUG] STDERR:\n", result.stderr)

        if result.returncode != 0:
            return jsonify({
                "success": False,
                "message": f"❌ Script failed:\n{result.stderr or result.stdout}"
            })

        for line in result.stdout.splitlines():
            if "beneficiaries" in line.lower():
                return jsonify({
                    "success": True,
                    "message": f"✅ {line.strip()}"
                })

        return jsonify({
            "success": True,
            "message": "✅ Sync completed"
        })

    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"❌ Error running sync: {e}"
        })

@app.route("/fsp-logout")
def fsp_logout():
    session.pop("fsp_logged_in", None)
    return redirect(url_for("fsp_login"))


@app.route("/scan")
def scan():
    # Only FSP-logged-in users should scan
    lang = request.args.get("lang", "en")

    if not session.get("fsp_logged_in"):
        return redirect(url_for("fsp_login", lang=lang))

    username = session.get("fsp_username", "User")
    program_id = session.get("fsp_program_id")

    # Fetch 121 program title
    program_title = ""
    if program_id:
        system_config = load_config()
        url121 = system_config.get("url121")
        if url121:
            try:
                login_resp = requests.post(
                    f"{url121}/api/users/login",
                    json={
                        "username": system_config.get("username121", ""),
                        "password": system_config.get("password121", "")
                    },
                    timeout=8
                )
                if login_resp.status_code == 201:
                    token = login_resp.json().get("access_token_general")
                    r = requests.get(
                        f"{url121}/api/programs/{program_id}",
                        cookies={"access_token_general": token},
                        timeout=8
                    )
                    if r.status_code == 200:
                        titles = r.json().get("titlePortal", {})
                        program_title = titles.get(lang) or next(iter(titles.values()), "")
            except Exception as e:
                print(f"[scan] Failed to fetch 121 program title: {e}")

    return render_template(
        "scan.html",
        lang=lang,
        t=translations.get(lang, translations["en"]),
        username=username,
        program_title=program_title,
        program_id=program_id or ""
    )


@app.route('/service-worker.js')
def sw():
    return send_from_directory('static', 'service-worker.js', mimetype='application/javascript')

@app.route('/manifest.webmanifest')
def manifest():
    return send_from_directory('static', 'manifest.webmanifest', mimetype='application/manifest+json')



from io import BytesIO

@app.route("/api/offline/latest.zip")
def api_offline_latest_zip():
    base_dir = "offline-cache"
    if not os.path.isdir(base_dir):
        return jsonify({"error": "No offline cache found"}), 404

    # Filter by programId if provided
    program_id_filter = request.args.get("programId")

    batch_dirs = []
    for d in os.listdir(base_dir):
        full_path = os.path.join(base_dir, d)
        if not os.path.isdir(full_path):
            continue
        # Check batch_info.json for programId match
        if program_id_filter:
            batch_info_path = os.path.join(full_path, "batch_info.json")
            if os.path.exists(batch_info_path):
                try:
                    with open(batch_info_path) as f:
                        batch_info = json.load(f)
                    if str(batch_info.get("programId")) != str(program_id_filter):
                        continue
                except Exception:
                    pass
        batch_dirs.append(full_path)

    if not batch_dirs:
        return jsonify({"error": "No batches found for this program"}), 404

    latest = max(batch_dirs, key=os.path.getmtime)

    # Zip the latest batch in memory
    mem = BytesIO()
    with zipfile.ZipFile(mem, "w", zipfile.ZIP_DEFLATED) as zf:
        for root, _, files in os.walk(latest):
            for fname in files:
                full_path = os.path.join(root, fname)
                arcname = os.path.relpath(full_path, latest)  # keep paths relative to batch root
                zf.write(full_path, arcname)
    mem.seek(0)

    return send_file(
        mem,
        mimetype="application/zip",
        as_attachment=True,
        download_name="latest_offline_cache.zip",
    )

@app.route('/ping')
def ping():
    return "ok", 200


@app.route("/beneficiary-offline")
def beneficiary_offline():
    # expected: /beneficiary-offline?uuid=<registrationReferenceId>&lang=en&program_id=<id>
    uuid = request.args.get("uuid")
    lang = request.args.get("lang", session.get("lang", "en"))
    session["lang"] = lang
    # CHANGE: don't return 400; render a shell so the SW can precache a 200
    if not uuid:
        uuid = ""

    # Prefer URL param, fall back to session (handles SW precache and fresh tabs)
    program_id = request.args.get("program_id") or session.get("fsp_program_id", "")

    # load display config scoped to the active program
    try:
        full_config = load_display_config()
        programs_map = full_config.get("programs", {}) or {}
        # Empty dict fallback (NOT the root config) — the root has no 'photo' key
        # so falling back to it would silently disable the photo section.
        program_config = programs_map.get(str(program_id), {}) if program_id else {}
        display_fields = program_config.get("fields", [])
        photo_config = program_config.get("photo", {})
    except Exception:
        display_fields = []
        photo_config = {}

    config = load_config()
    enc_key = config.get("ENCRYPTION_KEY", "")
    column_to_match = get_column_to_match(program_id) or config.get("COLUMN_TO_MATCH", "")

    return render_template(
        "beneficiary_offline.html",
        uuid=uuid,  # may be "", the page will prefer URL ?uuid=...
        lang=lang,
        t=translations.get(lang, translations["en"]),
        display_fields=display_fields,
        photo_config=photo_config,
        fernet_key=enc_key,
        column_to_match=column_to_match,
        program_id=program_id,
        program_currency=config.get("programCurrency", ""),
    )

@app.route("/success-offline")
def success_offline():
    lang = request.args.get("lang", "en")
    t = translations.get(lang, translations["en"])
    program_id = request.args.get("program_id") or session.get("fsp_program_id", "")
    return render_template("success_offline.html", lang=lang, t=t, program_id=program_id)

@app.route("/system-config.json")
def system_config_json():
    config = load_config()
    column = config.get("COLUMN_TO_MATCH")
    if not column:
        return jsonify({"error": "COLUMN_TO_MATCH missing"}), 500

    return jsonify({"COLUMN_TO_MATCH": column})



def get_column_to_match(program_id):
    """Fetch columnToMatch for a program from the 121 API.
    Falls back to system_config COLUMN_TO_MATCH if API unavailable."""
    config = load_config()
    url121 = config.get("url121")

    if url121 and program_id:
        try:
            token = get_121_token()
            if token:
                r = requests.get(
                    f"{url121}/api/programs/{program_id}/fsp-configurations",
                    cookies={"access_token_general": token},
                    timeout=10
                )
                if r.status_code == 200:
                    for fsp in r.json():
                        for prop in fsp.get("properties", []):
                            if prop.get("name") == "columnToMatch":
                                return prop.get("value")
        except Exception as e:
            print(f"[get_column_to_match] API error: {e}")

    # Fallback to per-program stored value, then legacy global value
    per_program = config.get("COLUMN_TO_MATCH_PER_PROGRAM", {})
    return per_program.get(str(program_id)) or config.get("COLUMN_TO_MATCH")


def get_121_token():
    import requests
    config = load_config()
    username = config.get("username121")
    password = config.get("password121")
    base_url = config.get("url121")

    if not username or not password or not base_url:
        print("❌ Missing 121 credentials")
        return None

    login_url = f"{base_url}/api/users/login"

    try:
        resp = requests.post(
            login_url,
            json={"username": username, "password": password},
            timeout=8
        )

        if resp.status_code != 201:
            print(f"❌ Login failed ({resp.status_code}): {resp.text}")
            return None

        # 121 API returns token in JSON (correct behaviour)
        token = resp.json().get("access_token_general")
        if not token:
            print("❌ Login succeeded but no token returned")
            return None

        return token

    except Exception as e:
        print(f"❌ 121 API error: {e}")
        return None


@app.route('/submit-payments', methods=['POST'])
def submit_payments():
    import csv
    import io
    import os
    import json
    from datetime import datetime
    from cryptography.fernet import Fernet

    # Load config
    config = load_config()
    # Use active program from session (multi-program support)
    program_id = session.get("fsp_program_id") or config.get("programId")
    fernet_key = config.get("ENCRYPTION_KEY")

    if not program_id:
        return "❌ No active program selected. Please go back and select a program.", 400

    # Fetch column_to_match from 121 API for this specific program
    column_to_match = get_column_to_match(program_id)

    if not column_to_match:
        return f"❌ Could not determine columnToMatch for program {program_id}. Check 121 FSP configuration.", 400

    if not fernet_key:
        return "❌ Missing ENCRYPTION_KEY in system_config.json", 400

    # Fernet decryptor
    try:
        fernet = Fernet(fernet_key.encode())
    except Exception as e:
        return f"❌ Invalid Fernet key: {e}", 400

    # Get uploaded CSV file
    if 'csv' not in request.files:
        return "❌ No CSV file provided", 400

    file = request.files['csv']
    if file.filename == '':
        return "❌ Empty filename", 400

    try:
        csv_content = file.stream.read().decode("utf-8")
    except Exception as e:
        return f"❌ Failed to read CSV: {e}", 400

    reader = csv.DictReader(io.StringIO(csv_content))
    rows = list(reader)

    if not rows:
        return "❌ CSV is empty", 400

    # -------------------------------
    # LOAD OFFLINE CACHE FOR PAYMENT MAPPING
    # -------------------------------
    cache_base = "offline-cache"

    import re
    def extract_batch_number(name):
        match = re.search(r"payment-recent-batch-(\d+)", name)
        return int(match.group(1)) if match else -1

    # Filter batch dirs to only those belonging to the active program
    all_dirs = [d for d in os.listdir(cache_base) if d.startswith("payment-recent-batch-")]
    program_dirs = []
    for d in all_dirs:
        batch_info_path = os.path.join(cache_base, d, "batch_info.json")
        if os.path.exists(batch_info_path):
            try:
                with open(batch_info_path) as f:
                    bi = json.load(f)
                if str(bi.get("programId")) == str(program_id):
                    program_dirs.append(d)
            except Exception:
                pass
        else:
            program_dirs.append(d)  # include legacy batches without batch_info

    batch_dirs = sorted(program_dirs, key=extract_batch_number)

    if not batch_dirs:
        return "❌ No recent payment batches found for this program — run sync first.", 400

    latest_batch = batch_dirs[-1]
    print(f"[DEBUG] Using batch folder: {latest_batch}")

    reg_cache_path = os.path.join(cache_base, latest_batch, "registrations_cache.json")
    if not os.path.exists(reg_cache_path):
        return "❌ registrations_cache.json missing — run sync again.", 400

    try:
        with open(reg_cache_path, "r", encoding="utf-8") as f:
            reg_data = json.load(f)
    except Exception as e:
        return f"❌ Failed to load registrations_cache.json — {e}", 500

    # -------------------------------
    # BUILD MAP: plaintext match column → paymentId
    # -------------------------------
    match_to_pid = {}

    for record in reg_data:
        uuid = record.get("uuid")
        payment_id = record.get("paymentId")

        encrypted_value = record.get("data", {}).get(column_to_match, "")

        if encrypted_value and payment_id:
            try:
                decrypted_value = fernet.decrypt(encrypted_value.encode()).decode().strip()
                match_to_pid[decrypted_value] = payment_id
            except Exception as e:
                print(f"[!] Failed to decrypt value for UUID {uuid}: {e}")

    # -------------------------------
    # GROUP CSV ROWS BY paymentId
    # -------------------------------
    grouped = {}

    for row in rows:
        raw_value = row.get(column_to_match, "").strip()
        status = row.get("status", "").strip()

        # If incoming value is still encrypted (rare)
        if raw_value.startswith("gAAAA"):
            try:
                raw_value = fernet.decrypt(raw_value.encode()).decode().strip()
            except Exception as e:
                print(f"[!] Failed to decrypt incoming {column_to_match}: {raw_value} — {e}")
                continue

        payment_id = match_to_pid.get(raw_value)

        if not payment_id:
            print(f"[!] No paymentId found for {column_to_match}: {raw_value}")
            continue

        grouped.setdefault(payment_id, []).append({
            column_to_match: raw_value,
            "status": status
        })

    if not grouped:
        return "❌ No valid rows to submit — check your CSV and sync data.", 400

    # -------------------------------
    # SUBMIT TO 121 /paymentId/excel-reconciliation
    # -------------------------------
    token = get_121_token()
    if not token:
        return "❌ Login to 121 failed", 401

    success_count = 0
    fail_count = 0

    for pid, items in grouped.items():
        output_buffer = io.StringIO()
        writer = csv.DictWriter(output_buffer, fieldnames=[column_to_match, "status"])
        writer.writeheader()

        for item in items:
            writer.writerow({
                column_to_match: item[column_to_match],
                "status": item["status"]
            })

        upload_url = f"{config['url121']}/api/programs/{program_id}/payments/{pid}/excel-reconciliation"

        files = {"file": ("reconciliation.csv", output_buffer.getvalue(), "text/csv")}

        upload_resp = requests.post(upload_url, files=files, cookies={"access_token_general": token})

        if upload_resp.status_code == 201:
            success_count += 1
            print(f"[OK] Submitted to paymentId {pid}")
        else:
            fail_count += 1
            print(f"[ERROR] Failed to submit to paymentId {pid}: {upload_resp.status_code} — {upload_resp.text}")

    # -------------------------------
    # FINAL RESPONSE
    # -------------------------------
    if success_count > 0:
        return f"✅ Submitted to {success_count} paymentId(s). ❌ {fail_count} failed.", 200
    else:
        return "❌ All submissions failed.", 500


@app.route("/invalid-qr")
def invalid_qr():
    # keep previously-selected language
    lang = request.args.get("lang", "en")
    reason = request.args.get("reason", "")

    program_id = request.args.get("program_id") or session.get("fsp_program_id", "")
    return render_template(
        "invalid-qr.html",
        reason=reason,
        lang=lang,
        t=translations.get(lang, translations["en"]),
        program_id=program_id
    )


@app.route("/vouchers", methods=["GET"])
def vouchers_page():
    if not session.get("admin_logged_in"):
        lang = request.args.get("lang", session.get("lang", "en"))
        return redirect(url_for("admin_login", lang=lang))

    # language
    lang = request.args.get("lang", session.get("lang", "en"))
    session["lang"] = lang
    t = translations.get(lang, translations["en"])

    # Load system config (for program title, currency, etc.)
    try:
        system_config = load_config()
    except Exception:
        system_config = {}

    # Username from session
    username = session.get("admin_username", "")

    # Program title from session OR fallback to system_config
    program_title = session.get("program_title") or system_config.get("programTitle", "")

    return render_template(
        "vouchers.html",
        lang=lang,
        t=t,
        program_title=program_title,
        username=username
    )



@app.route("/vouchers/upload", methods=["POST"])
def vouchers_upload():
    if not session.get("admin_logged_in"):
        return jsonify({"success": False, "message": "Not authorized"}), 403

    if "csv" not in request.files:
        return jsonify({"success": False, "message": "No file uploaded"}), 400

    f = request.files["csv"]
    filename = f.filename.lower()

    try:
        os.makedirs("uploads", exist_ok=True)
        upload_path = os.path.join("uploads", filename)
        f.save(upload_path)

        rows = []

        # ---------------------------------------------------------------------
        # CASE 1: CSV
        # ---------------------------------------------------------------------
        if filename.endswith(".csv"):
            with open(upload_path, "r", encoding="utf-8", errors="replace") as infile:
                reader = csv.DictReader(infile)

                for row in reader:
                    clean_row = {}
                    for key, value in row.items():
                        if key is None:
                            continue
                        clean_key = key.strip().replace("\ufeff", "").lower()
                        clean_row[clean_key] = value.strip() if isinstance(value, str) else value

                    ref = (
                        clean_row.get("referenceid")
                        or clean_row.get("reference id")
                        or clean_row.get("reference_id")
                        or clean_row.get("refid")
                        or clean_row.get("id")
                        or ""
                    )

                    clean_row["referenceid"] = ref.strip()
                    rows.append(clean_row)

        # ---------------------------------------------------------------------
        # CASE 2: XLSX
        # ---------------------------------------------------------------------
        elif filename.endswith(".xlsx"):
            from openpyxl import load_workbook
            wb = load_workbook(upload_path, read_only=True, data_only=True)
            ws = wb.active

            # Read header
            header_row = next(ws.iter_rows(values_only=True))
            headers = [str(h).strip().lower() if h else None for h in header_row]

            # Read remaining rows (skip header)
            is_header = True
            for row in ws.iter_rows(values_only=True):
                if is_header:
                    is_header = False
                    continue

                clean_row = {}
                for key, value in zip(headers, row):
                    if not key:
                        continue
                    clean_row[key] = value.strip() if isinstance(value, str) else value

                ref = (
                    clean_row.get("referenceid")
                    or clean_row.get("reference id")
                    or clean_row.get("reference_id")
                    or clean_row.get("refid")
                    or clean_row.get("id")
                    or ""
                )

                clean_row["referenceid"] = str(ref).strip()

                # Skip empty rows (all None)
                if any(v not in (None, "") for v in clean_row.values()):
                    rows.append(clean_row)

        # ---------------------------------------------------------------------
        # CASE 3: XLS
        # ---------------------------------------------------------------------
        elif filename.endswith(".xls"):
            import xlrd
            wb = xlrd.open_workbook(upload_path)
            sh = wb.sheet_by_index(0)

            headers = [str(h).strip().lower() for h in sh.row_values(0)]

            for rx in range(1, sh.nrows):  # start at row 1 (skip header)
                values = sh.row_values(rx)
                clean_row = {}

                for key, value in zip(headers, values):
                    clean_row[key] = value.strip() if isinstance(value, str) else value

                ref = (
                    clean_row.get("referenceid")
                    or clean_row.get("reference id")
                    or clean_row.get("reference_id")
                    or clean_row.get("refid")
                    or clean_row.get("id")
                    or ""
                )

                clean_row["referenceid"] = str(ref).strip()

                # Skip empty rows
                if any(v not in (None, "") for v in clean_row.values()):
                    rows.append(clean_row)

        else:
            return jsonify({"success": False, "message": "Unsupported file type"}), 400

        # Save metadata
        session["voucher_file_path"] = upload_path
        session["voucher_count"] = len(rows)

        return jsonify({"success": True, "count": len(rows)})

    except Exception as e:
        return jsonify({"success": False, "message": f"Failed to parse file: {e}"}), 400

@app.route("/vouchers/download", methods=["GET"])
def vouchers_download():
    if not session.get("admin_logged_in"):
        return redirect(url_for("admin_login", lang=request.args.get("lang", "en")))

    file_path = session.get("voucher_file_path")
    if not file_path or not os.path.exists(file_path):
        flash("No uploaded data to generate vouchers.", "error")
        return redirect(url_for("vouchers_page"))

    rows = []
    filename = file_path.lower()

    # -----------------------------
    # CASE 1: CSV
    # -----------------------------
    if filename.endswith(".csv"):
        with open(file_path, "r", encoding="utf-8", errors="replace") as infile:
            reader = csv.DictReader(infile.read().splitlines())
            for row in reader:
                clean_row = {k.strip().lower(): (v.strip() if isinstance(v, str) else v)
                             for k, v in row.items() if k}

                ref = (
                    clean_row.get("referenceid")
                    or clean_row.get("reference id")
                    or clean_row.get("reference_id")
                    or clean_row.get("refid")
                    or clean_row.get("id")
                    or ""
                )
                clean_row["referenceid"] = str(ref).strip()
                rows.append(clean_row)

    # -----------------------------
    # CASE 2: XLSX
    # -----------------------------
    elif filename.endswith(".xlsx"):
        from openpyxl import load_workbook

        wb = load_workbook(file_path, read_only=True)
        ws = wb.active

        rows_iter = ws.iter_rows(values_only=True)

        # --- Read REAL header row only once ---
        header_row = next(rows_iter)
        headers = [str(h).strip().lower() if h else "" for h in header_row]

        # --- Process all remaining rows (skips header correctly) ---
        for row in rows_iter:
            # skip empty rows
            if not any(row):
                continue

            clean_row = {}
            for key, value in zip(headers, row):
                if key:
                    clean_row[key] = (
                        str(value).strip() if isinstance(value, str) else value
                    )

            ref = (
                clean_row.get("referenceid")
                or clean_row.get("reference id")
                or clean_row.get("reference_id")
                or clean_row.get("refid")
                or clean_row.get("id")
                or ""
            )
            clean_row["referenceid"] = str(ref).strip()
            rows.append(clean_row)


    # -----------------------------
    # CASE 3: XLS
    # -----------------------------
    elif filename.endswith(".xls"):
        import xlrd
        wb = xlrd.open_workbook(upload_path)
        sh = wb.sheet_by_index(0)

        # --- Find the FIRST non-empty row → this is the real header row ---
        header_row_index = None
        for i in range(sh.nrows):
            row = sh.row_values(i)
            if any(str(c).strip() for c in row):   # Row contains at least 1 non-empty cell
                header_row_index = i
                break

        if header_row_index is None:
            return jsonify({"success": False, "message": "No valid header row found"}), 400

        # Extract header names
        headers = [
            str(h).strip().lower() if h is not None else ""
            for h in sh.row_values(header_row_index)
        ]

        # --- Process all rows AFTER the header row ---
        for rx in range(header_row_index + 1, sh.nrows):
            values = sh.row_values(rx)

            # skip empty rows
            if not any(str(v).strip() for v in values):
                continue

            clean_row = {}
            for key, value in zip(headers, values):
                if key:  # skip empty header columns
                    clean_row[key] = (
                        str(value).strip() if isinstance(value, str) else value
                    )

            # normalize referenceid
            ref = (
                clean_row.get("referenceid")
                or clean_row.get("reference id")
                or clean_row.get("reference_id")
                or clean_row.get("refid")
                or clean_row.get("id")
                or ""
            )
            clean_row["referenceid"] = str(ref).strip()

            rows.append(clean_row)

    else:
        flash("Unsupported voucher file type.", "error")
        return redirect(url_for("vouchers_page"))

    # -----------------------------
    # Generate PDF
    # -----------------------------
    pdf_io = generate_vouchers_pdf(
        rows,
        static_folder=os.path.join(app.root_path, "static")
    )

    return send_file(
        pdf_io,
        mimetype="application/pdf",
        as_attachment=True,
        download_name="vouchers.pdf"
    )